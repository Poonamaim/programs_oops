import openpyxl

wb = openpyxl.load_workbook('student.xlsx')
ws = wb.active
for j in range(1,ws.max_row+1):
    values = [ws.cell(row=j,column=i).value for i in range(1,ws.max_column+1)]
    Result = int(values[0]) + int(values[1]) + int(values[2]) + int(values[3])
  # result1 = (str(values[0]))+"|"+str(values[1])+"|"+str(values[2])+"|"+str(values[3])
result2 =Result/4
print(result2)




