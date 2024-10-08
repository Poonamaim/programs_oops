import openpyxl
import json

wb = openpyxl.load_workbook('student.xlsx')
ws = wb.active
for j in range(2, ws.max_row+1):
    values = [ws.cell(row=j, column=i).value for i in range(1, ws.max_column+1)]
    totalMarks = int(values[0])+int(values[1])+int(values[2])+int(values[3])
    percentage = totalMarks/4
    row = (str(values[0])) + "|" + str(values[1]) + "|" + str(values[2]) + "|" + str(values[3]) + "|" + str(totalMarks) + "|" + str(percentage)
    print(row)
    x = [i for i in row]
    print(x)
   
