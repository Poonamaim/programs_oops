import openpyxl

wb = openpyxl.load_workbook('student.xlsx')
ws = wb.active
for j in range(1,ws.max_row+1):
    data = [ws.cell(row=i, column=2).value for i in range(8, 12)]
    print(data)


