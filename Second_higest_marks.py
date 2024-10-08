import openpyxl

wb = openpyxl.load_workbook('student.xlsx')
ws = wb.active
totalMarksList = []
for j in range(2, ws.max_row+1):

    values = [ws.cell(row=j, column=i).value for i in range(1, ws.max_column+1)]

    totalMarks = int(values[0])+int(values[1])+int(values[2])+int(values[3])
    totalMarksList.append(totalMarks)
    # percentage = totalMarks/4
    # row = (str(values[0])) + "|" + str(values[1]) + "|" + str(values[2]) + "|" + str(values[3]) + "|"+str(totalMarks) + "|" + str(percentage)

print(totalMarksList)
totalMarksList.sort(reverse = True)
print(totalMarksList)
secondHighest = totalMarksList[2]
print(totalMarksList)
print(secondHighest)
exit()
  #   max1 = 0
  #   max2 = 0
  #   index = 0
  #   index2 = 0
  #   for i in range(len(row)):
  #     if int(row[i]) > max1:
  #       max2 = max
  #       max1 = row[i]
  #       index2 =index
  #       index = i
  #     elif row[i] >max2:
  #       max2 = row[i]
  #       index2 =i
  #     print(max2,index2)
  #

emp_id = ['E12','E56','E76','E78']
emp_details = [{"fname":"Govind", "lname":"Talde"},{"fname":"Poonam", "lname":"ghosh"},{"fname":"sayali", "lname":"balkawade"},{"fname":"komal", "lname":"thorat"}]


list =[
    {"id":"E12","Fname":"govind","lname":"talde"},
    {"id": "E56", "Fname": "poonam", "lname": "ghosh"},
    {"id": "E12", "Fname": "govind", "lname": "talde"},
    {"id": "E56", "Fname": "poonam", "lname": "ghosh"},

]
for i in emp