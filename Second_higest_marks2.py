import openpyxl

wb = openpyxl.load_workbook('student.xlsx')
ws = wb.active
totalMarksList = []
for j in range(1, ws.max_row+1):

    values = [ws.cell(row=j, column=i).value for i in range(1, ws.max_column+1)]
#  print(values)
    ist1 = []
    list1 = sorted(values, reverse=True)
    #print(list1)
    print(list1[0:13])





