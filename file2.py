import openpyxl

wb = openpyxl.load_workbook('Book1.xlsx')
ws = wb.active

for j in range(1,ws.max_row+1):
    values = [ws.cell(row=j,column=i).value for i in range(1,ws.max_column+1)]
    print(values)
