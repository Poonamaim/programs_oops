import openpyxl

wb = openpyxl.load_workbook('student.xlsx')
ws = wb.active
for j in range(2, ws.max_row+1):
    Values = [ws.cell(row=j, column=i).value for i in range(1, ws.max_column+1)]


    Result1 = int(Values[0])+int(Values[1])+int(Values[2])+int(Values[3])


    print(" total_marks"+"and "+ "percentage" )
    print(Result1)
   # Result = result1
    per = (Result1/4)
    print(per)


